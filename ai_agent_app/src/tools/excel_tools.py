"""
Excel automation tools.
Uses openpyxl for file-based generation and COM automation for live Excel control.
"""
import json
import os
from contextlib import contextmanager

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

try:
    import pythoncom
    from win32com.client import DispatchEx, GetActiveObject
except ImportError:  # pragma: no cover - non-Windows fallback
    pythoncom = None
    DispatchEx = None
    GetActiveObject = None


def _style_header_row(ws, num_cols: int):
    """Apply professional header styling."""
    header_fill = PatternFill(start_color="1A2B5E", end_color="1A2B5E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _style_data_row(ws, row_num: int, num_cols: int):
    """Alternating row colors for readability."""
    fill_color = "EBF0FA" if row_num % 2 == 0 else "FFFFFF"
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


@contextmanager
def _com_context():
    if pythoncom is None:
        yield
        return

    pythoncom.CoInitialize()
    try:
        yield
    finally:
        pythoncom.CoUninitialize()


def _require_excel_com():
    if pythoncom is None or DispatchEx is None or GetActiveObject is None:
        raise RuntimeError("Excel desktop automation requires Windows with pywin32 and Microsoft Excel installed.")


def _normalize_path(filename: str) -> str:
    return os.path.abspath(os.path.expanduser(filename))


def _is_active_workbook_target(filename: str = None) -> bool:
    return not filename or str(filename).strip().upper() in {"__ACTIVE__", "ACTIVE"}


def _excel_color_to_bgr(value: str) -> int:
    cleaned = value.strip().replace("#", "")
    if len(cleaned) != 6:
        raise ValueError("Color must be a 6-digit hex value like FF0000.")
    red = int(cleaned[0:2], 16)
    green = int(cleaned[2:4], 16)
    blue = int(cleaned[4:6], 16)
    return blue << 16 | green << 8 | red


def _style_live_excel_header(worksheet, col_count: int):
    """Apply simple professional styling to the first row in a live Excel sheet."""
    if col_count <= 0:
        return
    header_range = worksheet.Range(f"A1:{get_column_letter(col_count)}1")
    header_range.Font.Bold = True
    header_range.Font.Color = _excel_color_to_bgr("FFFFFF")
    header_range.Interior.Color = _excel_color_to_bgr("1A2B5E")
    header_range.HorizontalAlignment = -4108  # xlCenter


def _resolve_border_style(border_style: str):
    if not border_style:
        return None
    normalized = str(border_style).strip().lower()
    style_map = {
        "thin": 1,
        "continuous": 1,
        "solid": 1,
        "dashed": -4115,
        "dot": -4118,
    }
    if normalized in style_map:
        return style_map[normalized]
    try:
        return int(border_style)
    except Exception:
        return 1


def _get_excel_app(visible: bool = True):
    _require_excel_com()
    try:
        app = GetActiveObject("Excel.Application")
    except Exception:
        app = DispatchEx("Excel.Application")
    app.Visible = visible
    app.DisplayAlerts = False
    return app


def _find_workbook(app, filename: str):
    normalized = _normalize_path(filename).lower()
    for workbook in app.Workbooks:
        try:
            if os.path.abspath(workbook.FullName).lower() == normalized:
                return workbook
        except Exception:
            continue
    return None


def _open_or_create_workbook(app, filename: str, create_if_missing: bool = True):
    if _is_active_workbook_target(filename):
        if app.Workbooks.Count > 0:
            return app.ActiveWorkbook, False
        return app.Workbooks.Add(), True

    normalized = _normalize_path(filename)
    os.makedirs(os.path.dirname(normalized), exist_ok=True)

    workbook = _find_workbook(app, normalized)
    if workbook:
        return workbook, False

    if os.path.exists(normalized):
        return app.Workbooks.Open(normalized), False

    if not create_if_missing:
        raise FileNotFoundError(f"Workbook not found: {normalized}")

    workbook = app.Workbooks.Add()
    workbook.SaveAs(normalized)
    return workbook, True


def _get_worksheet(workbook, sheet_name: str = None, create_if_missing: bool = True):
    if not sheet_name:
        return workbook.ActiveSheet

    for sheet in workbook.Worksheets:
        if sheet.Name == sheet_name:
            return sheet

    if not create_if_missing:
        raise ValueError(f"Worksheet '{sheet_name}' not found.")

    worksheet = workbook.Worksheets.Add()
    worksheet.Name = sheet_name
    return worksheet


def _parse_matrix(values_json: str):
    try:
        values = json.loads(values_json)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid values_json: {exc}") from exc

    if not isinstance(values, list):
        raise ValueError("values_json must decode to a list.")

    matrix = []
    for row in values:
        if isinstance(row, list):
            matrix.append(row)
        else:
            matrix.append([row])
    return matrix


def _workbook_label(workbook, filename: str = None) -> str:
    return workbook.Name if _is_active_workbook_target(filename) else workbook.FullName


def _open_created_workbook_in_excel(filename: str, sheet_name: str = None) -> str:
    """Open a saved workbook in Excel when desktop automation is available."""
    try:
        with _com_context():
            app = _get_excel_app(visible=True)
            workbook, _ = _open_or_create_workbook(app, filename, create_if_missing=False)
            worksheet = _get_worksheet(workbook, sheet_name, create_if_missing=True)
            worksheet.Activate()
            workbook.Activate()
            return f" Opened in Excel: {workbook.FullName} (sheet: {worksheet.Name})"
    except Exception:
        try:
            os.startfile(_normalize_path(filename))
            return f" Opened in Excel using the default Windows app: {_normalize_path(filename)}"
        except Exception:
            return ""


def excel_create_table(filename: str, sheet_name: str = "Sheet1",
                       headers: list = None, rows: list = None) -> str:
    """Create a new Excel file with a formatted table and open it in Excel when possible."""
    try:
        if _is_active_workbook_target(filename):
            matrix = [headers or ["Column A", "Column B", "Column C"]]
            if rows:
                matrix.extend(rows)

            write_result = excel_write_range(
                filename="__ACTIVE__",
                sheet_name=sheet_name,
                start_cell="A1",
                values_json=json.dumps(matrix),
                autofit=True,
            )
            format_result = excel_format_range(
                filename="__ACTIVE__",
                sheet_name=sheet_name,
                range_address=f"A1:{get_column_letter(len(matrix[0]))}1",
                bold=True,
                autofit=True,
            )
            with _com_context():
                app = _get_excel_app(visible=True)
                workbook, _ = _open_or_create_workbook(app, "__ACTIVE__", create_if_missing=True)
                worksheet = _get_worksheet(workbook, sheet_name, create_if_missing=True)
                _style_live_excel_header(worksheet, len(matrix[0]))
            if write_result.startswith("❌"):
                return write_result
            return f"✅ Excel table created in the active workbook on sheet {sheet_name}. {format_result}"

        os.makedirs(os.path.dirname(os.path.abspath(filename)), exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not headers:
            headers = ["Column A", "Column B", "Column C"]

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        _style_header_row(ws, len(headers))

        if rows:
            for row_idx, row_data in enumerate(rows, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                _style_data_row(ws, row_idx, len(headers))

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 14)

        ws.row_dimensions[1].height = 24
        wb.save(filename)
        opened_message = _open_created_workbook_in_excel(filename, sheet_name)
        return f"✅ Excel table created: {filename} with {len(headers)} columns, {len(rows or [])} rows.{opened_message}"
    except Exception as e:
        return f"❌ Error creating Excel table: {e}"


def excel_insert_value(filename: str, row: int, column: int, value,
                       sheet_name: str = None) -> str:
    """Insert a value into a specific cell."""
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        ws.cell(row=row, column=column, value=value)
        wb.save(filename)
        col_letter = get_column_letter(column)
        return f"✅ Inserted '{value}' at cell {col_letter}{row} in {filename}"
    except Exception as e:
        return f"❌ Error inserting value: {e}"


def excel_fill_demo(filename: str) -> str:
    """Create a demo file with values 1-6 and open it in Excel when possible."""
    try:
        os.makedirs(os.path.dirname(os.path.abspath(filename)), exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Demo"

        headers = ["ID", "Value", "Category", "Score", "Rank", "Status"]
        demo_data = [
            [1, 100, "Alpha", 95.5, 1, "Active"],
            [2, 200, "Beta", 87.3, 2, "Active"],
            [3, 300, "Gamma", 76.1, 3, "Pending"],
            [4, 400, "Delta", 65.8, 4, "Active"],
            [5, 500, "Epsilon", 54.2, 5, "Inactive"],
            [6, 600, "Zeta", 43.9, 6, "Pending"],
        ]

        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        _style_header_row(ws, len(headers))

        for row_idx, row_data in enumerate(demo_data, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data_row(ws, row_idx, len(headers))

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 14)

        ws.row_dimensions[1].height = 24
        wb.save(filename)
        opened_message = _open_created_workbook_in_excel(filename, "Demo")
        return f"✅ Demo Excel created: {filename} with formatted values 1-6 example data.{opened_message}"
    except Exception as e:
        return f"❌ Error creating demo Excel: {e}"


def excel_open_workbook(filename: str, sheet_name: str = None,
                        visible: bool = True, create_if_missing: bool = True) -> str:
    """Open Excel desktop app and open or create a workbook."""
    try:
        with _com_context():
            app = _get_excel_app(visible=visible)
            workbook, created = _open_or_create_workbook(app, filename, create_if_missing=create_if_missing)
            worksheet = _get_worksheet(workbook, sheet_name, create_if_missing=True)
            worksheet.Activate()
            workbook.Activate()
            if not _is_active_workbook_target(filename):
                workbook.Save()
            action = "created and opened" if created else "opened"
            return f"✅ Excel workbook {action}: {_workbook_label(workbook, filename)} (sheet: {worksheet.Name})"
    except Exception as e:
        if not _is_active_workbook_target(filename):
            try:
                os.startfile(_normalize_path(filename))
                return f"✅ Opened workbook using the default Windows app: {_normalize_path(filename)}"
            except Exception:
                pass
        return f"❌ Error opening Excel workbook: {e}"


def excel_describe_workbook(filename: str, sheet_name: str = None) -> str:
    """Describe sheets and used range for a workbook."""
    try:
        with _com_context():
            app = _get_excel_app(visible=True)
            workbook, _ = _open_or_create_workbook(app, filename, create_if_missing=not _is_active_workbook_target(filename))
            worksheet = _get_worksheet(workbook, sheet_name, create_if_missing=False)
            used_range = worksheet.UsedRange
            row_count = int(used_range.Rows.Count or 0)
            col_count = int(used_range.Columns.Count or 0)
            sheet_names = ", ".join(sheet.Name for sheet in workbook.Worksheets)
            return (
                f"✅ Workbook: {_workbook_label(workbook, filename)}\n"
                f"Sheet: {worksheet.Name}\n"
                f"All sheets: {sheet_names}\n"
                f"Used range: {row_count} row(s) x {col_count} column(s)"
            )
    except Exception as e:
        return f"❌ Error describing workbook: {e}"


def excel_write_range(filename: str, start_cell: str, values_json: str,
                      sheet_name: str = None, autofit: bool = True) -> str:
    """Write a 2D array of values into a workbook open in Excel."""
    try:
        matrix = _parse_matrix(values_json)
        if not matrix:
            return "⚠️ No values provided to write."

        with _com_context():
            app = _get_excel_app(visible=True)
            workbook, _ = _open_or_create_workbook(app, filename, create_if_missing=True)
            worksheet = _get_worksheet(workbook, sheet_name, create_if_missing=True)

            row_count = len(matrix)
            col_count = max(len(row) for row in matrix)
            padded = tuple(tuple(row + [""] * (col_count - len(row))) for row in matrix)
            col_letters, start_row = coordinate_from_string(start_cell.upper())
            start_col = column_index_from_string(col_letters)

            # Cell-by-cell COM writes are slower but much more reliable for active workbooks.
            for row_idx, row_values in enumerate(padded, start=0):
                for col_idx, value in enumerate(row_values, start=0):
                    worksheet.Cells(start_row + row_idx, start_col + col_idx).Value = value

            if autofit:
                try:
                    end_row = start_row + row_count - 1
                    end_col = start_col + col_count - 1
                    target_range = worksheet.Range(
                        worksheet.Cells(start_row, start_col),
                        worksheet.Cells(end_row, end_col),
                    )
                    target_range.EntireColumn.AutoFit()
                except Exception:
                    pass

            worksheet.Activate()
            if not _is_active_workbook_target(filename):
                workbook.Save()
            return (
                f"✅ Wrote {row_count} row(s) x {col_count} column(s) to "
                f"{worksheet.Name}!{start_cell.upper()} in {_workbook_label(workbook, filename)}"
            )
    except Exception as e:
        return f"❌ Error writing Excel range: {e}"


def excel_read_range(filename: str, range_address: str, sheet_name: str = None) -> str:
    """Read values from a range in a workbook open in Excel."""
    try:
        with _com_context():
            app = _get_excel_app(visible=True)
            workbook, _ = _open_or_create_workbook(app, filename, create_if_missing=not _is_active_workbook_target(filename))
            worksheet = _get_worksheet(workbook, sheet_name, create_if_missing=False)
            values = worksheet.Range(range_address.upper()).Value
            if values is None:
                rendered = "empty"
            elif isinstance(values, tuple):
                rendered = json.dumps(
                    [list(row) if isinstance(row, tuple) else [row] for row in values],
                    default=str,
                )
            else:
                rendered = json.dumps([[values]], default=str)
            return f"✅ Values in {worksheet.Name}!{range_address.upper()}: {rendered}"
    except Exception as e:
        return f"❌ Error reading Excel range: {e}"


def excel_format_range(filename: str, range_address: str, sheet_name: str = None,
                       bold: bool = False, italic: bool = False,
                       autofit: bool = False, font_size: int = 0,
                       font_color: str = "", fill_color: str = "",
                       number_format: str = "", border_style: str = "",
                       horizontal_alignment: str = "", row_height: float = 0) -> str:
    """Apply simple formatting to a live Excel range."""
    try:
        with _com_context():
            app = _get_excel_app(visible=True)
            workbook, _ = _open_or_create_workbook(app, filename, create_if_missing=not _is_active_workbook_target(filename))
            worksheet = _get_worksheet(workbook, sheet_name, create_if_missing=False)
            cell_range = worksheet.Range(range_address.upper())

            if bold:
                cell_range.Font.Bold = True
            if italic:
                cell_range.Font.Italic = True
            if font_size:
                cell_range.Font.Size = font_size
            if font_color:
                cell_range.Font.Color = _excel_color_to_bgr(font_color)
            if fill_color:
                cell_range.Interior.Color = _excel_color_to_bgr(fill_color)
            if number_format:
                cell_range.NumberFormat = number_format
            if border_style:
                resolved_border_style = _resolve_border_style(border_style)
                for border_id in range(7, 13):
                    cell_range.Borders(border_id).LineStyle = resolved_border_style
            if horizontal_alignment:
                alignment_map = {
                    "left": -4131,
                    "center": -4108,
                    "right": -4152,
                }
                resolved_alignment = alignment_map.get(str(horizontal_alignment).strip().lower())
                if resolved_alignment is not None:
                    cell_range.HorizontalAlignment = resolved_alignment
            if row_height:
                cell_range.RowHeight = row_height
            if autofit:
                try:
                    cell_range.EntireColumn.AutoFit()
                except Exception:
                    pass

            if not _is_active_workbook_target(filename):
                workbook.Save()
            return f"✅ Formatted range {worksheet.Name}!{range_address.upper()} in {_workbook_label(workbook, filename)}"
    except Exception as e:
        return f"❌ Error formatting Excel range: {e}"


def excel_save_workbook(filename: str) -> str:
    """Save an open workbook."""
    try:
        with _com_context():
            app = _get_excel_app(visible=True)
            workbook, _ = _open_or_create_workbook(app, filename, create_if_missing=not _is_active_workbook_target(filename))
            workbook.Save()
            return f"✅ Saved Excel workbook: {_workbook_label(workbook, filename)}"
    except Exception as e:
        return f"❌ Error saving Excel workbook: {e}"


def excel_close_workbook(filename: str, save_changes: bool = True) -> str:
    """Close a workbook while leaving Excel available."""
    try:
        with _com_context():
            app = _get_excel_app(visible=True)
            workbook, _ = _open_or_create_workbook(app, filename, create_if_missing=not _is_active_workbook_target(filename))
            workbook_name = _workbook_label(workbook, filename)
            workbook.Close(SaveChanges=save_changes)
            return f"✅ Closed Excel workbook: {workbook_name}"
    except Exception as e:
        return f"❌ Error closing Excel workbook: {e}"
